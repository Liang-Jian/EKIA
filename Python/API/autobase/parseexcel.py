import openpyxl
from openpyxl.styles import Border, Side, Font
# from openpyxl import load_workbook
# from openpyxl import *
import time
# from UI.config.VarConfig import *
import os


class ParseExcel(object):

    def __init__(self):
        self.workbook = None
        self.excelFile = None
        # 设置字体颜色
        self.font = Font(color=None)
        # 颜色对应的RGB值
        self.RGBDict = {
            "red":"FFFF3030",
            "green":"FF008B00"
        }

    def load_work_book(self, excel_path_and_name):
        '''
        将Excel文件加载到内存，并获取其workbook对象
        '''
        try:
            # self.workbook = openpyxl.load_workbook(excel_path_and_name, data_only = True)
            self.workbook = openpyxl.load_workbook(excel_path_and_name)
        except Exception as e:
            raise e
        self.excelFile = excel_path_and_name
        return  self.workbook

    def get_sheet_by_name(self, sheetName):
        '''
        根据sheet名获取该sheet对象
        '''
        try:
            sheet = self.workbook.get_sheet_by_name(sheetName)
            # sheet = self.workbook[sheetName]
            return sheet
        except Exception as e:
            raise e

    def get_sheet_by_index(self, sheetIndex):
        '''
        根据sheet的索引获取该sheet对象
        '''
        try:
            sheetname = self.workbook.get_sheet_names()[sheetIndex]
        except Exception as e:
            raise e
        sheet = self.workbook.get_sheet_by_name(sheetname)
        return sheet

    def get_rows_number(self, sheet):
        '''
        获取sheet中有数据区域的结束行号
        '''
        return sheet.max_row

    def get_cols_number(self, sheet):
        '''
        获取sheet中有数据区域的结束列号
        '''
        return sheet.max_column

    def get_start_row_number(self, sheet):
        '''
        获取sheet中有数据区域的开始行号
        '''
        return sheet.min_row

    def get_start_col_number(self, sheet):
        '''
        获取sheet中有数据区域的开始列号
        '''
        return sheet.min_colum

    def get_row(self, sheet, rowNo):
        '''
        获取sheet中某一行，返回的是这一行所有的数据内容组成的tuple,下标从1开始，sheet.row[1]表示第一行
        '''
        try:
            # return sheet.rows[rowNo - 1]
            return list(sheet.rows[rowNo - 1])
        except Exception as e:
            raise e

    def get_column(self, sheet, rolNo):
        '''
        获取sheet中某一列，返回的是这一列所有的数据内容组成的tuple,下标从1开始，sheet.columns[1]表示第一列
        '''
        try:
            # return sheet.columns[rolNo - 1]
            return list(sheet.columns[rolNo - 1])
        except Exception as e:
            raise e


    def get_cell_of_value(self, sheet, coordinate = None, rowNo = None, colsNo = None):
        '''
        根据单元格所在的位置索引获取该单元格中的值，下标从1开始，sheet.cell(row = 1, column = 1).value，表示excel中第一行第一列的值
        '''
        if coordinate != None:
            try:
                return sheet.cell(coordinate = coordinate).value
            except Exception as e:
                raise e
        elif coordinate is None and rowNo is not None and colsNo is not None:
            try:
                return sheet.cell(row = rowNo, column = colsNo).value
            except Exception as e:
                raise e
        else:
            pass

    def get_cell_of_object(self, sheet, coordinate = None, rowNo = None, colsNo = None):
        '''
        获取某个单元格的对象，可以根据单元格所在的位置的数字索引，也可以直接根据Excel中单元格的编码及坐标，
        如get_cell_of_object(sheet, coordinate = "A1") or get_cell_of_object(sheet, rowNo = 1, colsNo = 2)
        '''
        if coordinate != None:
            try:
                return sheet.cell(coordinate = coordinate)
            except Exception as e:
                raise e
        elif coordinate == None and rowNo is not None and colsNo is not None:
            try:
                return sheet.cell(row = rowNo, column = colsNo)
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell !")


    def write_cell(self, sheet, content, coordinate = None, rowNo = None, colsNo = None, style = None):
        '''
        根据单元格在Excel中的编码坐标或者数字索引坐标向单元格中写入数据，
        下标从1开始，参数style表示字体的颜色的名字，比如red，green
        '''
        if coordinate is not None:
            try:
                sheet.cell(coordinate = coordinate).value = content
                if style is not None:
                    sheet.cell(coordinate = coordinate).font = Font(color = self.RGBDict[style], size=10)
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        elif coordinate == None and rowNo is not None and colsNo is not None:
            try:
                sheet.cell(row = rowNo, column = colsNo).value = content
                if style:
                    sheet.cell(row = rowNo, column = colsNo).font = Font(color = self.RGBDict[style], size=10)
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell !")

    def write_cell_current_time(self, sheet, coordinate = None, rowNo = None, colsNo = None, style = None):
        '''
        写入当前的时间，下标从1开始
        '''
        now = int(time.time()) # 显示时间戳
        time_array = time.localtime(now)
        current_time = time.strftime("%Y-%m-%d %H:%M:%S", time_array)
        if coordinate is not None:
            try:
                sheet.cell(coordinate = coordinate).value = current_time
                if style is not None:
                    sheet.cell(coordinate = coordinate).font = Font(color = self.RGBDict[style], size=10)
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        elif coordinate == None and rowNo is not None and colsNo is not None:
            try:
                sheet.cell(row = rowNo, column = colsNo).value = current_time
                if style:
                    sheet.cell(row = rowNo, column = colsNo).font = Font(color = self.RGBDict[style], size=10)
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell !")














if __name__ == '__main__':
    # 测试代码
    pe = ParseExcel()
    # 测试所需Excel文件“通用投保验证.xls”，请自行创建
    # path = parentDirPath + "\\testData\\通用投保验证_S20180290.xlsx"
    path = r"D:\zdhlog\EKIA\Python\API\autodata\testcase\冰鉴接口案例.xlsx"
    # path = "D:\\MyDocuments\\itw_liuyh01\桌面\\buikdNumber.xls"
    # print(path)
    pe.load_work_book(path)
    # print("通过名称获取sheet对象的名字：", pe.get_sheet_by_name("200").title)
    # print("通过名称获取sheet对象的名字：", pe.get_sheet_by_name("data"))
    # print("通过index序号获取sheet对象的名字：", pe.get_sheet_by_index(1).title)
    sheet = pe.get_sheet_by_name("出单")
    # sheet = pe.get_sheet_by_index(0)
    # print(sheet)
    # sheet = pe.get_sheet_by_index(0)
    # print("获取最大行号：", pe.get_rows_number(sheet)) # 获取最大行号
    # print("获取最大列号：", pe.get_cols_number(sheet)) # 获取最大列号
    # rows_num = pe.get_rows_number(sheet)
    # print(rows_num)

    # rows1 = pe.get_row(sheet, 9)
    # print(rows1)
    # rows2 = pe.get_row(sheet, 3)
    #

    # for i in range(2, rows1):

    # for i in range(2, 502):


    # for i in range(2, 202):
    #     rows1 = pe.get_row(sheet, i)
    #     # print(rows1)
    #     key = []
    #     # print("获取第一行：", rows1)
    #     for x in rows1:
    #         x = x.value
    #         key.append(x)
    #
    #     # print(key)
    #
    #     path = "/home/robot/ApiTestFrame/autodata/testcase/200.txt"
    #     f = open(path, "r")
    #     str = f.read()
    #
    #     str = str.replace("{{personnelName}}", key[0])
    #     # str = str.replace("{{sexCode}}", key[1])
    #     # str = str.replace("{{certificateType}}", key[2])
    #     str = str.replace("{{certificateNo}}", key[1])
    #     # str = str.replace("{{birthday}}", key[4])
    #
    #
    #     print(str)







    #
    # value = []
    # print("获取第二行：", rows2)
    # for y in rows2:
    #     y = y.value
    #     value.append(y)
    # print(value)
    #
    # z = dict(zip(key, value))
    # print(z)
    #
    #
    #
    # cols = pe.get_column(sheet, 1)
    # print("获取第一列：", cols)
    # for i in cols:
    #     print(i.value)
    #
    #
    #
    # # 获取第n行第n列单元格内容
    print("获取第一行第一列单元格内容：", pe.get_cell_of_value(sheet, rowNo=8, colsNo=1))
    # print("获取第一行第一列单元格内容：", pe.get_cell_of_value(sheet, coordinate="B3"))
    #
    # # 获取第n行第n列单元格对象
    # print("获取第一行第一列单元格对象：", pe.get_cell_of_object(sheet, rowNo=1, colsNo=1))
    # print("获取第一行第一列单元格对象：", pe.get_cell_of_object(sheet, coordinate="B3"))
    # print("获取第一行第一列单元格对象：", pe.get_cell_of_object(sheet, coordinate="B3").value)
    #
    # # pe.write_cell(sheet, "TK", rowNo=2, colsNo=9, style="red")
    # # pe.write_cell_current_time(sheet, rowNo=2, colsNo=10)
    # #
    # # pe.write_cell(sheet, "TK", rowNo=3, colsNo=9, style="green")
    # # pe.write_cell_current_time(sheet, rowNo=3, colsNo=10)
